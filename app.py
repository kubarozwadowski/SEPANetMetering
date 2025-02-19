import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import numpy as np
import plotly.express as px
import geopandas as gpd
from openpyxl import load_workbook
from monthly import calculate_stats_month, calculate_stats_uptomonth, calculate_stats_uptomonth_state, calculate_stats_range_months_state, calculate_stats_range_us_total
from monthlyapi import technology_year, technology_range, state_technology_range, state_technology_year
from calendar import month_abbr
from datetime import datetime
import plotly.graph_objects as go

#wb = load_workbook('net_metering2024.xlsx')
#monthly_totals_states = wb['Monthly_Totals-States']
#monthly_totals_us = wb['Monthly_Totals-US']

wb = load_workbook('NetMeteringAll.xlsx')
monthly_totals_states = wb['Sheet1']

st.set_page_config(page_title="US Net Metering Atlas", page_icon=":bar_chart:")


categories_mapping = {
    "Photovoltaic": {
        "Capacity MW": "E",
        "Customers": "J",
        "Virtual Capacity MW": "O",
        "Virtual Customers": "T",
        "Energy Sold Back MWh": "Y"
    },
    "Wind": {
        "Capacity MW": "BH",
        "Customers": "BM",
        "Energy Sold Back MWh": "BR"
    },
    "Battery": {
        "PV-Paired Battery Capacity (MW)": "AD",
        "PV-Paired Installations": "AI",
        "PV-Paired Energy Capacity (MWh)": "AN",
        "Not PV-Paired Battery Capacity (MW)": "AS",
        "PV-Paired Installations": "AX",
        "PV-Paired Energy Capacity (MWh)": "BC"
    },
    "Other": {
        "Capacity MW": "BW",
        "Customers": "CB",
        "Energy Sold Back MWh": "CG"
    },
    "All Technologies": {
        "Capacity MW": "CL",
        "Customers": "CQ",
        "Energy Sold Back MWh": "CV"
    }
}

month_mapping = {
    "Jan": 1,
    "Feb": 2,
    "Mar": 3,
    "Apr": 4,
    "May": 5,
    "Jun": 6,
    "Jul": 7,
    "Aug": 8,
    "Sep": 9,
    "Oct": 10,
    "Nov": 11,
    "Dec": 12
}
#for displaying month name in tables
month_reverse_mapping = {v: k for k, v in month_mapping.items()}

function_mapping = {
    "Calculate statistics for a specific month (US Total)" : calculate_stats_month,
    "Calculate statistics up to and including a specific month (US Total)": calculate_stats_uptomonth,
    "Calculate statistics by state up to and including a specific month": calculate_stats_uptomonth_state,
    "Calculate statistics by state for a range of months": calculate_stats_range_months_state,
    "Calculate statistics for a range of months (US Total)" : calculate_stats_range_us_total
}


infolink = "https://www.eia.gov/electricity/data/eia861m/#:~:text=Description%3A%20The,established%20imputation%20procedures."
faq = "https://www.eia.gov/survey/form/eia_861m/faqs.php"
st.logo("SEPALogo.png")
st.title("US Net Metering Atlas")
st.subheader("Interactive Analysis of Monthly Net Metering Data Across Various Technologies and States")

st.write("""
The data presented on this website is sourced from the U.S. Energy Information Administration (EIA). The EIA collects data through the Form EIA-861M, a monthly survey, and the EIA API for annual data. The primary focus is on net metering connections, capturing the communication and technological advances in the electricity sector. This data encompasses a comprehensive analysis of monthly and annual electricity usage by sector, categorized by technology and industry. This survey aims to provide insights into the growth and distribution of net metering customers and the energy they contribute back to the grid.
""")

with st.expander("Click to see important definitions and explanations of the categories:"):
    st.markdown('''
    - **Total:** residential + commercial + industrial + transportation + other.
    - **All Technologies:** photovoltaic + battery + wind + other. In 2022 and prior, battery data is omitted.
    - **Standard voltaic** is included in **Capacity MW** and **Customers** categories.
    - **Capacity MW:** Total installed capacity of PV systems.
    - **Customers:** Number of customers enrolled in net metering programs. Each meter is counted as a separate customer, even if multiple meters are aggregated under a single buyer representative for commercial, franchise, or residential groups. For public-street and highway lighting, count one customer per community.
    - **Virtual Customers:** Number of customers participating in virtual net metering programs.
    - **Energy Sold Back MWh:** Electricity generated by standard and virtual PV systems and the rate at which it is sold back to the grid.
    - **PV-Paired Battery Capacity (MW)** includes battery and photovoltaic systems.
    ''')

st.subheader("[More Information](%s)" % infolink)
st.subheader("[Frequently Asked Questions](%s)" % faq)


st.subheader("Search Features:")
st.write("""
- **Monthly Data Analysis:** Drawn from the Form EIA-861M, this feature allows users to perform a comparative analysis of electricity usage by sector and technology on a monthly basis. Users can search for specific metrics such as sellback rates and customer growth within different sectors and technologies.
- **Annual Data Analysis:** Utilizing the EIA API, this feature offers an annual overview of electricity usage and net metering statistics. Users can explore industry trends and sector-specific data to understand the broader patterns in net metering connections.
""")


st.subheader(":red[Important Notes About Data Discrepancies:]")
with st.expander("Due to the varied formatting of the data, some values may be missing, especially for older months. Click to see details."):
    st.write("""
    - Because of the nature of how the data is collected, monthly data over a large range may be slow to display graphs.
    - If a graph or data is not being displayed, the most likely cause is missing data.
    - No virtual data between Jan 2011 and Dec 2016.
    - No data for Tennessee in Jan 2011.
    - No data for Alabama and Mississippi between Jan 2011 and March 2012.
    """)

st.write("The data on this website is current as of **August 2024**.")


st.sidebar.title("User Input")
with st.sidebar:

    function = st.selectbox("Select a Function", [
        "Calculate statistics for a specific month (US Total)", 
        "Calculate statistics for a range of months (US Total)", 
        "Calculate statistics by state for a range of months",
        "Calculate statistics for a specific year (US Total)", 
        "Calculate statistics for a range of years (US Total)", 
        "Calculate statistics by state for a specific year", 
        "Calculate statistics by state for a range of years"
    ],index=None)

    if function == "Calculate statistics for a specific month (US Total)":
        with st.sidebar.form("form1"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            subcategory = st.selectbox('Select a Sector', ["Capacity MW", "Customers", "Energy Sold Back MWh"], index=None)
            with st.expander('Please Select a Month and Year'):
                this_year = datetime.now().year
                this_month = datetime.now().month
                year = st.selectbox("", range(2024, 2011, -1))
                month_abbr = month_abbr[1:]
                report_month_str = st.radio("", month_abbr, index=this_month - 1, horizontal=True)
                month = month_abbr.index(report_month_str) + 1
            submit_button = st.form_submit_button("Calculate")

    if function == "Calculate statistics for a range of months (US Total)":
        with st.sidebar.form("form2"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            subcategory = st.selectbox('Select a Sector', ["Capacity MW", "Customers", "Energy Sold Back MWh"], index=None)
            with st.expander('Please Select a Start Month and Year'):
                this_year = datetime.now().year
                this_month = datetime.now().month
                start_year = st.selectbox("", range(2024, 2011, -1), key="start_year")
                start_month_abbr = list(month_abbr)[1:]
                start_report_month_str = st.radio("", start_month_abbr, index=this_month - 1, horizontal=True, key="start_month")
                start_month = start_month_abbr.index(start_report_month_str) + 1
            with st.expander('Please Select an End Month and Year'):
                this_year = datetime.now().year
                this_month = datetime.now().month
                end_year = st.selectbox("", range(2024, 2011, -1), key="end_year")
                end_month_abbr = list(month_abbr)[1:]
                end_report_month_str = st.radio("", end_month_abbr, index=this_month - 1, horizontal=True, key="end_month")
                end_month = end_month_abbr.index(end_report_month_str) + 1
            submit_button = st.form_submit_button("Calculate")

    if function == "Calculate statistics by state for a range of months":
        with st.sidebar.form("form3"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            subcategory = st.selectbox('Select a Sector', ["Capacity MW", "Customers", "Energy Sold Back MWh"], index=None)
            with st.expander('Please Select a Start Month and Year'):
                this_year = datetime.now().year
                this_month = datetime.now().month
                start_year = st.selectbox("", range(2024, 2011, -1), key="start_year")
                start_month_abbr = list(month_abbr)[1:]
                start_report_month_str = st.radio("", start_month_abbr, index=this_month - 1, horizontal=True, key="start_month")
                start_month = start_month_abbr.index(start_report_month_str) + 1
            with st.expander('Please Select an End Month and Year'):
                this_year = datetime.now().year
                this_month = datetime.now().month
                end_year = st.selectbox("", range(2024, 2011, -1), key="end_year")
                end_month_abbr = list(month_abbr)[1:]
                end_report_month_str = st.radio("", end_month_abbr, index=this_month - 1, horizontal=True, key="end_month")
                end_month = end_month_abbr.index(end_report_month_str) + 1
            state = st.selectbox('Select a State', [
                "AK", "AL", "AR", "AZ", "CA", "CO", "CT", "DE", "FL", "GA",
                "HI", "IA", "ID", "IL", "IN", "KS", "KY", "LA", "MA", "MD",
                "ME", "MI", "MN", "MO", "MS", "MT", "NC", "ND", "NE", "NH",
                "NJ", "NM", "NV", "NY", "OH", "OK", "OR", "PA", "RI", "SC",
                "SD", "TN", "TX", "UT", "VA", "VT", "WA", "WI", "WV", "WY"], index=None)
            submit_button = st.form_submit_button("Calculate")

    if function == "Calculate statistics for a specific year (US Total)":
        with st.sidebar.form("form4"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            year = st.slider("Select a Year", 2013, 2019, 2016)
            submit_button = st.form_submit_button("Calculate")

    if function == "Calculate statistics for a range of years (US Total)":
        with st.sidebar.form("form5"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            values = st.slider("Select a range of values", 2013, 2019, (2014, 2017))   
            start_year = values[0]
            end_year = values[1]
            submit_button = st.form_submit_button("Calculate")
    
    if function == "Calculate statistics by state for a specific year":
        with st.sidebar.form("form6"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            year = st.slider("Select a Year", 2013, 2019, 2015)
            state = st.selectbox('Select a State', [
                 "AK", "AL", "AR", "AZ", "CA", "CO", "CT", "DE", "FL", "GA",
                 "HI", "IA", "ID", "IL", "IN", "KS", "KY", "LA", "MA", "MD",
                 "ME", "MI", "MN", "MO", "MS", "MT", "NC", "ND", "NE", "NH",
                 "NJ", "NM", "NV", "NY", "OH", "OK", "OR", "PA", "RI", "SC",
                 "SD", "TN", "TX", "UT", "VA", "VT", "WA", "WI", "WV", "WY"], index=None)
            submit_button = st.form_submit_button("Calculate")
    
    if function == "Calculate statistics by state for a range of years":
        with st.sidebar.form("form7"):
            category = st.selectbox('Select a Technology', ["Photovoltaic", "Wind", "Other", "All Technologies"], index=None)
            values = st.slider("Select a range of values", 2013, 2019, (2015, 2017))   
            start_year = values[0]
            end_year = values[1]
            state = st.selectbox('Select a State', [
                 "AK", "AL", "AR", "AZ", "CA", "CO", "CT", "DE", "FL", "GA",
                 "HI", "IA", "ID", "IL", "IN", "KS", "KY", "LA", "MA", "MD",
                 "ME", "MI", "MN", "MO", "MS", "MT", "NC", "ND", "NE", "NH",
                 "NJ", "NM", "NV", "NY", "OH", "OK", "OR", "PA", "RI", "SC",
                 "SD", "TN", "TX", "UT", "VA", "VT", "WA", "WI", "WV", "WY"], index=None)
            submit_button = st.form_submit_button("Calculate")
    


#show stats + maps + graphs on button press "Save"

if function and submit_button:
        

        if function == "Calculate statistics for a specific month (US Total)":
            month_num = month_mapping[report_month_str]
            start_column = categories_mapping[category][subcategory]
            stats_us, raw_data = calculate_stats_month(monthly_totals_states, year, month_num, start_column)

            if stats_us:
                st.write(f"## Statistics for {year} - {report_month_str}")

            
                plot_data = {
                    'State': [],
                    'Residential': [],
                    'Commercial': [],
                    'Industrial': [],
                    'Transportation': [],
                    'Total': []
                }

                for state_data in raw_data:
                    state = state_data['state']
                    category = state_data['category']
                    value = state_data['value']
                    
                    if state not in plot_data['State']:
                        plot_data['State'].append(state)
                        for cat in ['Residential', 'Commercial', 'Industrial', 'Total']:
                            plot_data[cat].append(0)

                    state_index = plot_data['State'].index(state)
                    plot_data[category][state_index] = value


                
                df_plot = pd.DataFrame(plot_data)


                # Convert any small floating-point values to zero explicitly
                df_plot = df_plot.applymap(lambda x: 0 if isinstance(x, (int, float)) and abs(x) <= 1e-5 else x)

                # Plot the data
                fig = go.Figure()
                for category in df_plot.columns[1:]:  # Skip the 'State' column
                    fig.add_trace(go.Bar(
                        x=df_plot['State'],
                        y=df_plot[category],
                        name=category
                    ))

                fig.update_layout(
                    title=f'Statistics for {year} - {report_month_str}',
                    xaxis_tickfont_size=14,
                    yaxis=dict(
                        title='Capacity (MW)',
                        titlefont_size=16,
                        tickfont_size=14,
                        range=[0, None]  # Set the y-axis to start at 0
                    ),
                    legend=dict(
                        x=0,
                        y=1.0,
                        bgcolor='rgba(255, 255, 255, 0)',
                        bordercolor='rgba(255, 255, 255, 0)'
                    ),
                    barmode='group',
                    bargap=0.15,
                    bargroupgap=0.1
                )

                st.plotly_chart(fig)

                # Display the data as a table
                st.write("## Tabular Data")
                st.dataframe(df_plot)

                # Create a mapping of state abbreviations to full state names
                state_abbreviation_mapping = {
                    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California",
                    "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia",
                    "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
                    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
                    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi", "MO": "Missouri",
                    "MT": "Montana", "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey",
                    "NM": "New Mexico", "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
                    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
                    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
                    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming"
                }

                # Plot choropleth maps for each sector
                st.write("## Choropleth Maps of State Data by Sector")
                if raw_data:
                    us_states = gpd.read_file('https://raw.githubusercontent.com/python-visualization/folium/master/examples/data/us-states.json')
                    us_states = us_states.to_crs(epsg=4326)  # Set CRS to WGS84 (EPSG:4326)

                    # Map state abbreviations to full names in raw_data
                    for entry in raw_data:
                        if entry['state'] in state_abbreviation_mapping:
                            entry['state'] = state_abbreviation_mapping[entry['state']]



                    for category in ['Residential', 'Commercial', 'Industrial', 'Total']:
                        st.write(f"### {category} {subcategory} by State")
                        df_category = pd.DataFrame([data for data in raw_data if data['category'] == category])

                        if not df_category.empty:
                            # Merge the GeoDataFrame with your raw data based on state names
                            gdf = us_states.merge(df_category, left_on='name', right_on='state', how='outer')
                            gdf.geometry = gdf.make_valid()  # Filter only valid geometries
                            gdf['value'] = pd.to_numeric(gdf['value'], errors='coerce').fillna(0).astype(float)

                            # Ensure no negative values exist
                            gdf['value'] = gdf['value'].apply(lambda x: 0 if x < 0 else x)

                            #debugging ; printing
                            print("Debug - Merged GeoDataFrame:\n", gdf[['name', 'value']])


                            # Plot the choropleth map
                            fig, ax = plt.subplots(1, 1, figsize=(20, 15))
                            gdf.plot(column='value', ax=ax, legend=True,
                                    cmap='coolwarm',  
                                    legend_kwds={'label': f"{category} {subcategory} by State",
                                                'orientation': "horizontal"},
                                    vmin=0, vmax=gdf['value'].max() * 1.1)  # Set minimum value for the color scale to ensure consistent interpretation
                            plt.title(f"{category} {subcategory}")
                            st.pyplot(fig)







        if function == "Calculate statistics by state for a range of months":
            start_column = categories_mapping[category][subcategory]

            # Call the calculate_stats_range_months_state function with the start and end year and month inputs
            stats, raw_data = calculate_stats_range_months_state(monthly_totals_states, start_year, end_year, start_month, end_month, state, start_column)
            
            if stats:
                st.write(f"## Statistics for {state} from {start_year}-{month_reverse_mapping[start_month]} to {end_year}-{month_reverse_mapping[end_month]}")
                data = {
                    'Category': [],
                    'Max Value': [],
                    'Min Value': [],
                    'Avg Value': []
                }

                for category, values in stats.items():
                    if values['max']['value'] is not None:
                        max_val_str = f"{values['max']['value']} ({values['max']['year']}-{month_reverse_mapping[values['max']['month']]})"
                    else:
                        max_val_str = "N/A"
                    
                    if values['min']['value'] is not None:
                        min_val_str = f"{values['min']['value']} ({values['min']['year']}-{month_reverse_mapping[values['min']['month']]})"
                    else:
                        min_val_str = "N/A"

                    data['Category'].append(category)
                    data['Max Value'].append(max_val_str)
                    data['Min Value'].append(min_val_str)
                    data['Avg Value'].append(values['avg'])

                df = pd.DataFrame(data)
                category_order = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
                df['Category'] = pd.Categorical(df['Category'], categories=category_order, ordered=True)
                st.dataframe(df)

        if function == "Calculate statistics for a range of months (US Total)":
            start_column = categories_mapping[category][subcategory]
            stats, raw_data = calculate_stats_range_us_total(monthly_totals_states, start_year, end_year, start_month, end_month, start_column)

            if stats:
                st.write("## Statistics")
                data = {
                    'Category': [],
                    'Max Value': [],
                    'Min Value': [],
                    'Avg Value': []
                }

                for category, values in stats.items():
                    data['Category'].append(category)
                    data['Max Value'].append(f"{values['max']['value']} ({values['max']['year']}-{month_reverse_mapping[values['max']['month']]})")
                    data['Min Value'].append(f"{values['min']['value']} ({values['min']['year']}-{month_reverse_mapping[values['min']['month']]})")
                    data['Avg Value'].append(values['avg'])

                df = pd.DataFrame(data)
                category_order = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
                df['Category'] = pd.Categorical(df['Category'], categories=category_order, ordered=True)
                st.dataframe(df)

                # Determine y-axis label based on subcategory
                y_axis_label = 'Value'
                if 'Capacity MW' in subcategory or 'Virtual Capacity MW' in subcategory:
                    y_axis_label = 'Capacity (MW)'
                elif 'Energy Sold Back MWh' in subcategory:
                    y_axis_label = 'Energy Sold Back (MWh)'

                # Create line graph for the range of months and years
                all_months_data = []
                for year in range(start_year, end_year + 1):
                    for month in range(1, 13):
                        if (year == start_year and month < start_month) or (year == end_year and month > end_month):
                            continue
                        monthly_stats, raw_month_data = calculate_stats_month(monthly_totals_states, year, month, start_column)
                        if monthly_stats:
                            for category, values in monthly_stats.items():
                                all_months_data.append({
                                    'Year': year,
                                    'Month': month,
                                    'Category': category,
                                    'Value': values.get('us_value')
                                })

                df_all_months = pd.DataFrame(all_months_data)
                df_all_months['Year-Month'] = df_all_months.apply(lambda row: f"{row['Year']}-{month_reverse_mapping[row['Month']]}", axis=1)
                fig = px.line(df_all_months, x='Year-Month', y='Value', color='Category', markers=True,
                              labels={'Value': y_axis_label})
                fig.update_xaxes(tickmode='array', tickvals=df_all_months['Year-Month'], ticktext=df_all_months['Year-Month'])
                st.plotly_chart(fig)

                # Create tabular data for the graph
                tabular_data = {
                    'Year-Month': df_all_months['Year-Month'].unique()
                }
                for category in category_order:
                    tabular_data[category] = df_all_months[df_all_months['Category'] == category]['Value'].values

                df_tabular = pd.DataFrame(tabular_data)
                st.write("## Tabular Data for Graph")
                st.dataframe(df_tabular)

        if function == "Calculate statistics for a specific year (US Total)":
            result = technology_year(category, year)
            if result:
                for sector, data in result['sectors'].items():
                    sector_name = data['data'][0]['sectorName'].capitalize()
                    sector_title = sector_name if sector_name != "All sectors" else "All Sectors"
                    st.write(f"### {sector_title} ({year})")
                    
                    # Extract US total capacity and customers
                    us_total_capacity = sum(float(item['capacity']) for item in data['data'] if item['state'] == 'US')
                    us_total_customers = sum(int(item['customers']) for item in data['data'] if item['state'] == 'US')
                    
                    summary_df = pd.DataFrame({
                        'Metric': ['Max Capacity', 'Min Capacity', 'Average Capacity', 
                                'Max Customers', 'Min Customers', 'Average Customers', 
                                'US Total Capacity', 'US Total Customers'],
                        'Value': [f"{data['max_capacity']['value']} ({data['max_capacity']['state']})", 
                                f"{data['min_capacity']['value']} ({data['min_capacity']['state']})", 
                                data['average_capacity'], 
                                f"{data['max_customers']['value']} ({data['max_customers']['state']})", 
                                f"{data['min_customers']['value']} ({data['min_customers']['state']})", 
                                data['average_customers'], 
                                f"{us_total_capacity} MW", 
                                f"{us_total_customers}"]
                    })
                    
                    st.dataframe(summary_df)
                    
                    # Create bar graphs for capacity and customers
                    detailed_df = pd.DataFrame(data['data'])
                    fig_capacity = px.bar(detailed_df, x='stateName', y='capacity', color='sector',
                                        title=f'Capacity by Sector for {year}', labels={'capacity': 'Capacity (MWh)'})
                    fig_customers = px.bar(detailed_df, x='stateName', y='customers', color='sector',
                                        title=f'Customers by Sector for {year}', labels={'customers': 'Number of Customers'})
                    
                    st.plotly_chart(fig_capacity)
                    st.plotly_chart(fig_customers)

        if function == "Calculate statistics for a range of years (US Total)":
            result = technology_range(category, start_year, end_year)
            if result:
                for sector, data in result['sectors'].items():
                    sector_name = data['data'][0]['sectorName'].capitalize()
                    sector_title = sector_name if sector_name != "All sectors" else "All Sectors"
                    st.write(f"### {sector_title} ({start_year} - {end_year})")
                    
                    # Extract yearly capacity and customer data
                    yearly_capacity = {item['period']: item['capacity'] for item in data['data'] if item['state'] == 'US'}
                    yearly_customers = {item['period']: item['customers'] for item in data['data'] if item['state'] == 'US'}
                    
                    # Prepare the summary table
                    summary_data = {
                        'Metric': ['Max Capacity', 'Min Capacity', 'Average Capacity', 
                                'Max Customers', 'Min Customers', 'Average Customers'],
                        'Value': [f"{data['max_capacity']['value']} ({data['max_capacity']['state']})", 
                                f"{data['min_capacity']['value']} ({data['min_capacity']['state']})", 
                                data['average_capacity'], 
                                f"{data['max_customers']['value']} ({data['max_customers']['state']})", 
                                f"{data['min_customers']['value']} ({data['min_customers']['state']})", 
                                data['average_customers']]
                    }
                    
                    # Add the current year values
                    for year in range(start_year, end_year + 1):
                        summary_data['Metric'].append(f"Capacity {year}")
                        summary_data['Value'].append(yearly_capacity.get(str(year), "N/A"))
                        summary_data['Metric'].append(f"Customers {year}")
                        summary_data['Value'].append(yearly_customers.get(str(year), "N/A"))
                    
                    summary_df = pd.DataFrame(summary_data)
                    
                    st.dataframe(summary_df)
                    
                    # Prepare data for trend graph
                    yearly_data = [item for item in data['data'] if item['state'] == 'US']
                    trend_df = pd.DataFrame(yearly_data)

                    # Ensure unique data points
                    trend_df = trend_df.drop_duplicates().dropna()

                    # Sort data by year
                    trend_df = trend_df.sort_values(by='period')

                    # Create line graphs for trends
                    fig_capacity_trend = px.line(trend_df, x='period', y='capacity', title='Total Capacity Trend',
                                                labels={'capacity': 'Capacity (MWh)', 'period': 'Year'})
                    fig_capacity_trend.update_xaxes(dtick=1)
                    fig_customers_trend = px.line(trend_df, x='period', y='customers', title='Total Customers Trend',
                                                labels={'customers': 'Number of Customers', 'period': 'Year'})
                    fig_customers_trend.update_xaxes(dtick=1)
                    
                    st.plotly_chart(fig_capacity_trend)
                    st.plotly_chart(fig_customers_trend)

        if function == "Calculate statistics by state for a specific year":
            result = state_technology_year(category, year, state)
            if result:
                for sector, data in result['sectors'].items():
                    sector_name = data['data'][0]['sectorName'].capitalize()
                    sector_title = sector_name if sector_name != "All sectors" else "All Sectors"
                    st.write(f"### {sector_title} ({state}, {year})")
                    
                    # Extract current capacity and customers for the specified year and state
                    current_capacity = next((item['capacity'] for item in data['data'] if item['period'] == str(year)), None)
                    current_customers = next((item['customers'] for item in data['data'] if item['period'] == str(year)), None)
                    
                    # Display current capacity and customers
                    summary_df = pd.DataFrame({
                        'Metric': ['Current Capacity', 'Current Customers'],
                        'Value': [f"{current_capacity} MW", f"{current_customers}"]
                    })
                    
                    st.dataframe(summary_df)
                    
                    # Create bar graphs for capacity and customers
                    detailed_df = pd.DataFrame(data['data'])
                    fig_capacity = px.bar(detailed_df, x='stateName', y='capacity', color='sector',
                                        title=f'Capacity by Sector for {state} ({year})', labels={'capacity': 'Capacity (MW)'})
                    fig_customers = px.bar(detailed_df, x='stateName', y='customers', color='sector',
                                        title=f'Customers by Sector for {state} ({year})', labels={'customers': 'Number of Customers'})
                    
                    st.plotly_chart(fig_capacity)
                    st.plotly_chart(fig_customers)

        if function == "Calculate statistics by state for a range of years":
            result = state_technology_range(category, start_year, end_year, state)
            if result:
                for sector, data in result['sectors'].items():
                    sector_name = data['data'][0]['sectorName'].capitalize()
                    sector_title = sector_name if sector_name != "All sectors" else "All Sectors"
                    st.write(f"### {sector_title} ({state}, {start_year} - {end_year})")

                    # Extract yearly capacity and customer data
                    yearly_capacity = {item['period']: item['capacity'] for item in data['data'] if item['state'] == state}
                    yearly_customers = {item['period']: item['customers'] for item in data['data'] if item['state'] == state}

                    # Prepare the summary table
                    summary_data = {
                        'Metric': ['Max Capacity', 'Min Capacity', 'Average Capacity', 
                                'Max Customers', 'Min Customers', 'Average Customers'],
                        'Value': [f"{data['max_capacity']['value']} ({data['max_capacity']['state']})", 
                                f"{data['min_capacity']['value']} ({data['min_capacity']['state']})", 
                                data['average_capacity'], 
                                f"{data['max_customers']['value']} ({data['max_customers']['state']})", 
                                f"{data['min_customers']['value']} ({data['min_customers']['state']})", 
                                data['average_customers']]
                    }

                    # Add the current year values
                    for year in range(start_year, end_year + 1):
                        summary_data['Metric'].append(f"Capacity {year}")
                        summary_data['Value'].append(yearly_capacity.get(str(year), "N/A"))
                        summary_data['Metric'].append(f"Customers {year}")
                        summary_data['Value'].append(yearly_customers.get(str(year), "N/A"))

                    summary_df = pd.DataFrame(summary_data)

                    st.dataframe(summary_df)

                    # Prepare data for trend graph
                    yearly_data = [item for item in data['data'] if item['state'] == state]
                    trend_df = pd.DataFrame(yearly_data)

                    # Ensure unique data points
                    trend_df = trend_df.drop_duplicates().dropna()

                    # Sort data by year
                    trend_df = trend_df.sort_values(by='period')

                    # Create line graphs for trends
                    fig_capacity_trend = px.line(trend_df, x='period', y='capacity', title='Total Capacity Trend',
                                                labels={'capacity': 'Capacity (MWh)', 'period': 'Year'})
                    fig_capacity_trend.update_xaxes(dtick=1)
                    fig_customers_trend = px.line(trend_df, x='period', y='customers', title='Total Customers Trend',
                                                labels={'customers': 'Number of Customers', 'period': 'Year'})
                    fig_customers_trend.update_xaxes(dtick=1)

                    st.plotly_chart(fig_capacity_trend)
                    st.plotly_chart(fig_customers_trend)