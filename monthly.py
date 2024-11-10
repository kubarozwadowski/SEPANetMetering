import openpyxl
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
from netMetering import col_to_index, is_valid_integer


wb = load_workbook('net_metering2024.xlsx')
monthly_totals_states = wb['Monthly_Totals-States']
monthly_totals_us = wb['Monthly_Totals-US']

# UhKbugjHqnLMfOetbnKlXUd1wVI42kvFjWV8Bdy6

def calculate_stats_month(sheet, year, month, start_column):
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[4:]  # Data starts on row 5

    # Filter by year, month, and state 'US'
    filtered_data = [row for row in data if row[0] == year and row[1] == month and row[2] == 'US']

    if not filtered_data:
        print(f"No data found for year {year}, month {month}, and state 'US'")
        return None, None

    # Define the categories in the specified order
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    start_idx = col_to_index(start_column) - 1  # Convert column letter to index (0-based)
    
    stats = {}
    raw_data = []

    for idx, category in enumerate(categories):
        column_index = start_idx + idx
        valid_entries = [(row[2], row[column_index]) for row in filtered_data if row[column_index] is not None]
        valid_entries = [(state, value) for state, value in valid_entries if is_valid_integer(value)]
        
        if valid_entries:
            us_value = next((value for state, value in valid_entries if state == 'US'), None)
            if us_value is not None:
                stats[category] = {
                    'us_value': us_value
                }
            raw_data.extend([{'state': state, 'category': category, 'value': value} for state, value in valid_entries])
        else:
            stats[category] = {
                'us_value': None
            }

    return stats, raw_data

def calculate_stats_uptomonth_state(sheet, month, state, start_column):
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[4:]  # Data starts on row 5

    # Filter by month and state, ensuring month is not None
    filtered_data = [row for row in data if row[1] is not None and row[1] <= month and row[2] == state]

    if not filtered_data:
        print(f"No data found for month {month} and state {state}")
        return None

    # Define the categories in the specified order
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    start_idx = col_to_index(start_column) - 1  # Convert column letter to index (0-based)
    
    stats = {}

    for idx, category in enumerate(categories):
        column_index = start_idx + idx
        valid_entries = [(row[1], row[2], row[column_index]) for row in filtered_data if row[column_index] is not None]
        valid_entries = [(month, state, value) for month, state, value in valid_entries if is_valid_integer(value)]
        
        if valid_entries:
            max_value = max(valid_entries, key=lambda x: x[2])
            min_value = min(valid_entries, key=lambda x: x[2])
            avg_value = sum(int(value) for _, _, value in valid_entries) / len(valid_entries)
            current_value = valid_entries[-1][2]  # Get the current month's value

            stats[category] = {
                'max': {'value': max_value[2], 'state': max_value[1], 'month': max_value[0]},
                'min': {'value': min_value[2], 'state': min_value[1], 'month': min_value[0]},
                'avg': avg_value,
                'current': current_value
            }
        else:
            stats[category] = {
                'max': {'value': None, 'state': None, 'month': None},
                'min': {'value': None, 'state': None, 'month': None},
                'avg': None,
                'current': None
            }

    return stats

def calculate_stats_uptomonth(sheet, month, start_column):
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[4:]  # Data starts on row 5

    # Filter by month, ensuring month is not None
    filtered_data = [row for row in data if row[1] is not None and row[1] <= month]

    if not filtered_data:
        print(f"No data found for month {month}")
        return None

    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    start_idx = col_to_index(start_column) - 1  # Convert column letter to index (0-based)
    
    stats = {}

    for idx, category in enumerate(categories):
        column_index = start_idx + idx
        valid_entries = [(row[1], row[2], row[column_index]) for row in filtered_data if row[column_index] is not None]
        valid_entries = [(month, state, value) for month, state, value in valid_entries if is_valid_integer(value)]
        
        if valid_entries:
            max_value = max(valid_entries, key=lambda x: x[2])
            min_value = min(valid_entries, key=lambda x: x[2])
            avg_value = sum(int(value) for _, _, value in valid_entries) / len(valid_entries)
            current_value = valid_entries[-1][2]  # Get the current month's value

            stats[category] = {
                'max': {'value': max_value[2], 'state': max_value[1], 'month': max_value[0]},
                'min': {'value': min_value[2], 'state': min_value[1], 'month': min_value[0]},
                'avg': avg_value,
                'current': current_value
            }
        else:
            stats[category] = {
                'max': {'value': None, 'state': None, 'month': None},
                'min': {'value': None, 'state': None, 'month': None},
                'avg': None,
                'current': None
            }

    return stats

def calculate_stats_range_months_state(sheet, start_year, end_year, start_month, end_month, state, start_column):
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[4:]  # Data starts on row 5

    # Define a helper function to check month and year ranges correctly
    def is_in_date_range(year, month):
        if start_year < end_year:
            # Standard range across multiple years
            return start_year <= year <= end_year and (
                (year == start_year and month >= start_month) or
                (year == end_year and month <= end_month) or
                (start_year < year < end_year)
            )
        elif start_year == end_year:
            # Within the same year
            return year == start_year and start_month <= month <= end_month
        else:
            # Wrap-around range (e.g., start_month=11, end_month=3)
            return (year == start_year and month >= start_month) or \
                   (year == end_year and month <= end_month)

    # Filter by state, year range, and month range
    filtered_data = [
        row for row in data
        if row[0] is not None and row[1] is not None and row[2] == state and is_in_date_range(row[0], row[1])
    ]

    if not filtered_data:
        print(f"No data found for state {state} from month {start_month} to {end_month} and year {start_year} to {end_year}")
        return None, None

    # Define the categories in the specified order
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    start_idx = col_to_index(start_column) - 1  # Convert column letter to index (0-based)
    
    stats = {}
    raw_data = []

    for idx, category in enumerate(categories):
        column_index = start_idx + idx
        valid_entries = [(row[0], row[1], row[column_index]) for row in filtered_data if row[column_index] is not None]
        valid_entries = [(year, month, value) for year, month, value in valid_entries if is_valid_integer(value)]
        
        if valid_entries:
            max_value = max(valid_entries, key=lambda x: x[2])
            min_value = min(valid_entries, key=lambda x: x[2])
            avg_value = sum(int(value) for year, month, value in valid_entries) / len(valid_entries)

            stats[category] = {
                'max': {'value': max_value[2], 'year': max_value[0], 'month': max_value[1]},
                'min': {'value': min_value[2], 'year': min_value[0], 'month': min_value[1]},
                'avg': avg_value,
                'data': valid_entries  # Add raw data for the graph
            }
            raw_data.extend([{'year': year, 'month': month, 'category': category, 'value': value} for year, month, value in valid_entries])
        else:
            stats[category] = {
                'max': {'value': None, 'year': None, 'month': None},
                'min': {'value': None, 'year': None, 'month': None},
                'avg': None,
                'data': []
            }

    return stats, raw_data


def calculate_stats_range_us_total(sheet, start_year, end_year, start_month, end_month, start_column):
    rows = list(sheet.iter_rows(values_only=True))
    data = rows[4:]  # Data starts on row 5

    # Define a helper function to check month and year ranges correctly
    def is_in_date_range(year, month):
        if start_year < end_year:
            # Standard range across multiple years
            return start_year <= year <= end_year and (
                (year == start_year and month >= start_month) or
                (year == end_year and month <= end_month) or
                (start_year < year < end_year)
            )
        elif start_year == end_year:
            # Within the same year
            return year == start_year and start_month <= month <= end_month
        else:
            # Wrap-around range (e.g., start_month=11, end_month=3)
            return (year == start_year and month >= start_month) or \
                   (year == end_year and month <= end_month)

    # Filter by US total, year range, and month range
    filtered_data = [
        row for row in data
        if row[0] is not None and row[1] is not None and row[2] == 'US' and is_in_date_range(row[0], row[1])
    ]

    if not filtered_data:
        print(f"No data found for US total from month {start_month} {start_year} to {end_month} {end_year}")
        return None, None

    # Define the categories in the specified order
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    start_idx = col_to_index(start_column) - 1  # Convert column letter to index (0-based)
    
    stats = {}
    raw_data = []

    for idx, category in enumerate(categories):
        column_index = start_idx + idx
        valid_entries = [(row[0], row[1], row[column_index]) for row in filtered_data if row[column_index] is not None]
        valid_entries = [(year, month, value) for year, month, value in valid_entries if is_valid_integer(value)]
        
        if valid_entries:
            max_value = max(valid_entries, key=lambda x: x[2])
            min_value = min(valid_entries, key=lambda x: x[2])
            avg_value = sum(int(value) for year, month, value in valid_entries) / len(valid_entries)

            stats[category] = {
                'max': {'value': max_value[2], 'year': max_value[0], 'month': max_value[1]},
                'min': {'value': min_value[2], 'year': min_value[0], 'month': min_value[1]},
                'avg': avg_value,
                'data': valid_entries  # Add raw data for the graph
            }
            raw_data.extend([{'year': year, 'month': month, 'category': category, 'value': value} for year, month, value in valid_entries])
        else:
            stats[category] = {
                'max': {'value': None, 'year': None, 'month': None},
                'min': {'value': None, 'year': None, 'month': None},
                'avg': None,
                'data': []
            }

    return stats, raw_data


