import openpyxl
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt

wb = load_workbook('net_metering2024.xlsx')
utility_level_states = wb['Utility_Level-States']
utility_level_territories = wb['Utility_Level-Ter']
utility_level_tpo = wb['Utility_Level-TPO']
monthly_totals_states = wb['Monthly_Totals-States']
monthly_totals_territories = wb['Monthly_Totals-Ter']
current_month_states = wb['Current_Month-States']
current_month_territories = wb['Current_Month-Ter']
monthly_totals_us = wb['Monthly_Totals-US']

def col_to_index(col):
    return openpyxl.utils.column_index_from_string(col)

def is_valid_integer(value):
    try:
        int(value)
        return True
    except (ValueError, TypeError):
        return False

all_technologies_capacitymw = []

#creates a dictionary per each subcategory in format {utility name, state, residential, commercial, industrial, transportation, total}
def create_dict(currentwb, listname, startcol):
    start_index = col_to_index(startcol)
    for row in currentwb.iter_rows(min_row=5, values_only=True):
        values = row[start_index - 1:start_index + 4]
        utility_data = {
            'Utility_Name': row[col_to_index('E') - 1],
            'State': row[col_to_index('C') - 1],
            'Residential': row[col_to_index('CO') - 1],
            'Commercial': row[col_to_index('CP') - 1],
            'Industrial': row[col_to_index('CQ') - 1],
            'Transportation': row[col_to_index('CR') - 1],
            'Total': row[col_to_index('CS') - 1]
        }
        listname.append(utility_data)

#return average, max, min of state. expects dictionary and state parameters
def list_by_state(listname, state):
    filtered_data = [entry for entry in listname if entry['State'] == state]
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    stats = {category: {'max': None, 'max_utility': '', 'min': None, 'min_utility': '', 'avg': None} for category in categories}

    for category in categories:
        valid_entries = [(entry['Utility_Name'], entry['State'], entry[category]) for entry in filtered_data if is_valid_integer(entry[category])]
        min_valid_entries = [entry for entry in valid_entries if 'Adjustment' not in entry[0]]

        if valid_entries:
            values = [int(entry[2]) for entry in valid_entries]
            avg_value = sum(values) / len(values)
            max_value = max(values)
            max_utility_entry = next(entry for entry in valid_entries if int(entry[2]) == max_value)
            max_utility = max_utility_entry[0]
            max_utility_state = max_utility_entry[1]

            if min_valid_entries:
                min_values = [int(entry[2]) for entry in min_valid_entries]
                min_value = min(min_values)
                min_utility_entry = next(entry for entry in min_valid_entries if int(entry[2]) == min_value)
                min_utility = min_utility_entry[0]
                min_utility_state = min_utility_entry[1]
            else:
                min_value = None
                min_utility = None
                min_utility_state = None

            stats[category]['max'] = max_value
            stats[category]['max_utility'] = (max_utility, max_utility_state)
            stats[category]['min'] = min_value
            stats[category]['min_utility'] = (min_utility, min_utility_state)
            stats[category]['avg'] = avg_value
        else:
            return "There are no entries for that state."
    return stats

#return max, min, and average of each topic, dictionary and state parameters
def list_stats(listname, category):
    valid_entries = [(entry['Utility_Name'], entry['State'], entry[category]) for entry in listname if is_valid_integer(entry[category])]
    min_valid_entries = [entry for entry in valid_entries if 'Adjustment' not in entry[0]]

    if not valid_entries:
        return "There are no matching entries."

    values = [int(entry[2]) for entry in valid_entries]
    avg_value = sum(values) / len(values)
    max_value = max(values)

    max_entry = next(entry for entry in valid_entries if int(entry[2]) == max_value)

    if min_valid_entries:
        min_values = [int(entry[2]) for entry in min_valid_entries]
        min_value = min(min_values)
        min_entry = next(entry for entry in min_valid_entries if int(entry[2]) == min_value)
    else:
        min_value = None
        min_entry = (None, None)

    return {
        'max': {'value': max_value, 'utility_name': max_entry[0], 'state': max_entry[1]},
        'min': {'value': min_value, 'utility_name': min_entry[0], 'state': min_entry[1]},
        'avg': float(avg_value)
    }

#compiles min, max, and avg of each category for each state into a dict
def list_stats_by_state(listname):
    states = set(entry['State'] for entry in listname)
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']
    stats = {state: {category: {} for category in categories} for state in states}

    for state in states:
        for category in categories:
            filtered_data = [entry for entry in listname if entry['State'] == state]
            stats[state][category] = list_stats(filtered_data, category)

    return stats

#plots values
def plot_stats(stats_by_state, subcategory):
    states = list(stats_by_state.keys())
    categories = ['Residential', 'Commercial', 'Industrial', 'Transportation', 'Total']

    for category in categories:
        max_values = []
        min_values = []
        avg_values = []
        valid_states = []
        
        for state in states:
            try:
                max_value = stats_by_state[state][category]['max']['value']
                min_value = stats_by_state[state][category]['min']['value']
                avg_value = stats_by_state[state][category]['avg']
                
                max_values.append(max_value)
                min_values.append(min_value)
                avg_values.append(avg_value)
                valid_states.append(state)
            except KeyError:
                continue
            except TypeError:
                continue

        plt.figure(figsize=(10, 6))
        x = range(len(valid_states))

        plt.plot(x, max_values, label='Max', marker='o')
        plt.plot(x, min_values, label='Min', marker='o')
        plt.plot(x, avg_values, label='Avg', marker='o')

        plt.xticks(x, valid_states, rotation='vertical')
        plt.xlabel('State')
        plt.ylabel(f'{subcategory} ({category})')
        plt.title(f'{subcategory} - {category}')
        plt.legend()
        plt.tight_layout()
        plt.show()

   


#HOW TO GET RAW DATA:
#step 1: create a dict where you specify worksheet and start column
#step 2: run second function based on specifities

#create_dict(utility_level_states, all_technologies_capacitymw, 'CO')
#print(list_stats(all_technologies_capacitymw, 'Residential'))


#HOW TO PLOT DATA:
#1: create a dictionary with create dict with inputs of worksheet, empty list, and start column
#2: Create new dictionary with data on each state with dictionary from step 1 as input
#3: plot using plot_stats using dict from step 2 and category (Capacity MW, Energy sold back, etc)

#create_dict(utility_level_states, all_technologies_capacitymw, 'CO')
#stats_by_state = list_stats_by_state(all_technologies_capacitymw)
#plot_stats(stats_by_state, 'Capacity MW')

#import pprint
#pprint.pprint(stats_by_state)